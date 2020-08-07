import openpyxl
from openpyxl.styles import PatternFill

workbook = openpyxl.load_workbook('./test.xlsx')
sheet = workbook.active

sheet['A1'] = 'Hello, World!'
sheet['A1'].fill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')

x = 1

if x > sheet['B4'].value:
    sheet['B4'] = x
    sheet['B4'].fill = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')

elif x < sheet['B4'].value:
    sheet['B4'] = x
    sheet['B4'].fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')

else:
    sheet['B4'] = x

workbook.save('./test.xlsx')