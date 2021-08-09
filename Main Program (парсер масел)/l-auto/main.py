import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill
requests.packages.urllib3.disable_warnings() 
import colorama
from colorama import Fore, Back, Style
colorama.init()

class Site():

    def __init__(self, HOST, URL):
        self.HOST = HOST
        self.URL = URL
        self.HEADERS = {
            'Accept': '*/*',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:78.0) Gecko/20100101 Firefox/78.0'
        }

    def get_html(self):
        r = requests.get(self.URL, headers=self.HEADERS, verify=False)
        return r

    def get_content(self, html, arg, select_class, find_str, location):
        soup = BeautifulSoup(html, 'html.parser')

        # delete old div price tag 
        for div in soup.find_all('div', {'class':'offer__product-old-price'}):
            div.decompose()
        
        cards = []
        
        try:
            testItem = eval('soup.' + find_str)
            cards.append(
                {
                    'price': str(testItem)
                }
            )
            print(Fore.WHITE + testItem + ' --- ' + self.URL)
            return cards
        except:
            text = 'Позиции больше нет ' + location
            print(Fore.RED + text)
            cards.append(
                {
                    'price': str(0)
                }
            )
            return cards
        

    def save_document(self, testItem, location):
        workbook = openpyxl.load_workbook('../Анализ.xlsx')
        sheet = workbook.active

        for item in testItem:
            if item['price'] != '':
                if item['price'] > str(sheet[location].value):
                    sheet[location] = item['price']
                    sheet[location].fill = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')

                elif item['price'] < str(sheet[location].value):
                    sheet[location] = item['price']
                    sheet[location].fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')


                else:
                    sheet[location] = item['price']
                    sheet[location].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

            else:
                sheet[location] = 'Нет товара'
                sheet[location].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        workbook.save('../Анализ.xlsx')

    def parser(self, arg, select_class, find_str, location):
        html = self.get_html()
        if html.status_code == 200:
            cards = []
            cards.extend(self.get_content(html.text, arg, select_class, find_str, location))
            self.save_document(cards, location)