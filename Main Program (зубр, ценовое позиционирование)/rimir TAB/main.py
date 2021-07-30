import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill
requests.packages.urllib3.disable_warnings() 

class Site():

    def __init__(self, HOST, URL):
        self.HOST = HOST
        self.URL = URL
        self.HEADERS = {
            'Accept': '*/*',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:78.0) Gecko/20100101 Firefox/78.0'
        }

    def get_html(self):
        r = requests.get(self.URL, headers=self.HEADERS, verify=False)
        return r

    def get_content(self, html, arg, select_class, find_str):
        soup = BeautifulSoup(html, 'html.parser')

        for sup in soup.find_all('sup', {}):
            sup.decompose()

        testItem = eval('soup.' + find_str)
        cards = []
        cards.append(
            {
                'price': testItem
            }
        )
        print(testItem + ' --- ' + self.URL)
        return cards

    def save_document(self, testItem, location):
        workbook = openpyxl.load_workbook('../Зубр.xlsx')
        sheet = workbook.active

        for item in testItem:
            if item['price'] != '':
                if item['price'] > str(sheet[location].value):
                    sheet[location] = item['price']
                    sheet[location].fill = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')

                elif item['price'] < str(sheet[location].value):
                    sheet[location] = item['price']
                    sheet[location].fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')


                else:
                    sheet[location] = item['price']
                    sheet[location].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

            else:
                sheet[location] = 'Нет товара'
                sheet[location].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        workbook.save('../Зубр.xlsx')

    def parser(self, arg, select_class, find_str, location):
        html = self.get_html()
        if html.status_code == 200:
            cards = []
            cards.extend(self.get_content(html.text, arg, select_class, find_str))
            self.save_document(cards, location)